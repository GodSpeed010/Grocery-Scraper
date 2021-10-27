import requests
import bs4
from bs4 import BeautifulSoup
import openpyxl

def main():
    get_wholefoods_sales()

def get_wholefoods_sales():
    sales_url = 'https://www.wholefoodsmarket.com/sales-flyer?store-id=10215' #John's Creek Whole Foods

    response_obj = requests.get(sales_url)
    
    #If there's an problem reaching the site, prints error msg
    try:
        response_obj.raise_for_status()
    except:
        print(f'Error reaching {sales_url}')
    
    html_content = BeautifulSoup(response_obj.text, 'html.parser')
    
    #Arrays of various sale items properties with html content
    sale_brand = html_content.find_all('div', class_="w-sales-tile__brand")    
    sale_title = html_content.find_all('h4', class_="w-sales-tile__product")
    sale_price = html_content.find_all('span', class_="w-sales-tile__sale-price w-header3 w-bold-txt")
    regular_price = html_content.find_all('div', class_="w-sales-tile__regular-price")
    unit_of_measurement = html_content.find_all('div', class_="w-sales-tile__uom")

    sale_properties = [sale_brand, sale_title, sale_price, regular_price, unit_of_measurement]
    sale_properties = to_text(sale_properties) #removes the html content to get only the clean data needed

    save_to_excel(sale_properties)

def to_text(arr):
    for x in range(len(arr)):
        for y in range(len(arr[x])):
            arr[x][y] = arr[x][y].text.strip()
            
            if arr[x][y].startswith('Regular'):
                arr[x][y] = arr[x][y].strip('Regular ')
    return arr

def print_data(arr):
    for col in range(len(arr[0])):
        for row in range(len(arr)):
            print(arr[row][col] + ' ')
        print('\n')

def save_to_excel(arr):
    wb = openpyxl.Workbook()
    ws = wb.active

    col_names = ['Brand', 'Product', 'Sale Price', 'Regular Price', 'Unit of measurement']
    #Write all column names to file
    for col, col_name in enumerate(col_names, start=1):
        ws.cell(row=1, column=col, value=col_name)

    #Write all sale data to file

    #Loop through columns
    for x, sale_property in enumerate(arr, start=1):

        #Loop through rows starting beneath col names
        for y, data in enumerate(sale_property, start=2):
            ws.cell(row=y, column=x, value=data)

    wb.save('output.xlsx')

main()